# update_xlsx_ia.py — añade sección IA Predictor al ESQUEMA + errores ERR-24/25/26
# Idempotente: se puede volver a ejecutar sin duplicar filas ni hojas.
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = 'ESQUEMA_SISTEMA_PORRA2026.xlsx'
DST = 'ESQUEMA_SISTEMA_PORRA2026.xlsx'

wb = load_workbook(SRC)

# ── Estilos (alineados con update_xlsx.py) ───────────────────────────────────
title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill('solid', start_color='1F4E78')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='2E75B6')
section_font = Font(name='Arial', size=12, bold=True)
section_fill = PatternFill('solid', start_color='D9E1F2')
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
border_thin = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)
highlight_fill = PatternFill('solid', start_color='E2EFDA')


# ── 1. Edge Functions: añadir porra-ia-compute v6 ────────────────────────────
ws_ef = wb['Edge Functions']
has_ia = any(
    'porra-ia-compute' in str(ws_ef.cell(row=r, column=1).value or '')
    for r in range(1, ws_ef.max_row + 1)
)
if not has_ia:
    nr = ws_ef.max_row + 1
    vals = [
        'porra-ia-compute v6',
        'Claude.ai via SQL (scrape_*) / pg_cron futuro (compute diario)',
        'Wikipedia MediaWiki API + 11v11.com HTML + Supabase DB (ia_* tablas)',
        'IA Predictor (Fases A-C implementadas). Router: status / scrape_elo (Wikipedia Module:SportsRankings → ia_elo_fifa) / scrape_h2h (11v11.com/stats → ia_h2h) / scrape_last5 (11v11.com/matches → ia_last5_results) / compute (Fase E pendiente). verify_jwt=false.',
        '✅ En producción (Fases A-C). Fase E/F pendientes.',
    ]
    for col, v in enumerate(vals, 1):
        c = ws_ef.cell(row=nr, column=col)
        c.value = v
        c.font = cell_font
        c.alignment = cell_align
        c.border = border_thin
        c.fill = highlight_fill


# ── 2. Sheet nueva: "IA Predictor" ───────────────────────────────────────────
IA_SHEET = 'IA Predictor'
if IA_SHEET in wb.sheetnames:
    del wb[IA_SHEET]
ws = wb.create_sheet(IA_SHEET)

ws['A1'] = 'PORRA MUNDIAL 2026 — IA Predictor (Fases A-F)'
ws.merge_cells('A1:D1')
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center_align

# ── 2a. Bloque: Arquitectura 3 capas ─────────────────────────────────────────
row = 3
ws.cell(row=row, column=1).value = 'Arquitectura 3 capas'
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = section_fill

row += 1
for col, h in enumerate(['Capa', 'Responsable', 'Entrada', 'Salida'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

layers = [
    ('1 — Ingesta', 'EF porra-ia-compute (4 actions scraper)',
     'Wikipedia Module:SportsRankings (ELO) + 11v11.com (H2H, últimos N)',
     'Tablas ia_elo_fifa, ia_h2h, ia_last5_results'),
    ('2 — Cómputo', 'EF porra-ia-compute action=compute (Fase E pendiente)',
     'ia_elo_fifa + ia_h2h + ia_last5_results + partido {home,away}',
     'ia_predictions (sign 1|X|2, confidence 0-100, breakdown JSONB)'),
    ('3 — Consumo', 'Frontend scoring.js / ko.js (Fase F pendiente)',
     'ia_predictions (RLS public read via policy ia_predictions_public_read)',
     'Tarjeta partido con pronóstico IA + bonus +1 pt si usuario opuesto y acierta'),
]
for cap, resp, inp, out in layers:
    row += 1
    for col, v in enumerate([cap, resp, inp, out], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin

# ── 2b. Bloque: Fórmula ──────────────────────────────────────────────────────
row += 2
ws.cell(row=row, column=1).value = 'Fórmula del pronóstico (Fase E)'
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = section_fill

row += 1
for col, h in enumerate(['Señal', 'Peso', 'Peso (fallback sin H2H)', 'Notas'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

formula = [
    ('ELO FIFA', '50%', '66%', 'ia_elo_fifa.elo_points. Fuente: Wikipedia Module:SportsRankings (incluye amistosos).'),
    ('H2H histórico', '25%', '0% (fallback)', 'ia_h2h. Fuente 11v11.com/stats (RSSSF-backed). Si el pair no tiene partido histórico, rebalancear.'),
    ('Racha últimos N', '25%', '34%', 'ia_last5_results.results (JSONB). N=8 default, ampliable a N=10 antes del 11 jun vía body.limit.'),
]
for signal, weight, fb, notes in formula:
    row += 1
    for col, v in enumerate([signal, weight, fb, notes], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin

row += 1
ws.cell(row=row, column=1).value = 'Umbrales signo 1|X|2 sobre raw_home_pct: >60% → 1 · 40-60% → X · <40% → 2.'
ws.cell(row=row, column=1).font = Font(name='Arial', size=10, italic=True)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


# ── 2c. Bloque: Tablas DB ────────────────────────────────────────────────────
row += 2
ws.cell(row=row, column=1).value = 'Tablas Supabase (migración 20260421_create_ia_predictor_tables.sql, Fase A)'
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = section_fill

row += 1
for col, h in enumerate(['Tabla', 'PK / índices', 'Columnas clave', 'Origen / RLS'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

tables = [
    ('ia_elo_fifa',
     'PK team_code (ISO-3). Index idx_ia_elo_fifa_scraped implícito.',
     'team_code, team_name, elo_points NUMERIC(7,2), rank_position, scraped_at, source',
     'Wikipedia Module:SportsRankings (Fase B.2). RLS enabled, sin policy pública (service role only).'),
    ('ia_last5_results',
     'PK team_code (FK → ia_elo_fifa.team_code ON DELETE CASCADE). Index idx_ia_last5_scraped (scraped_at DESC).',
     'team_code, results JSONB (array N objects {date, opponent_name, opponent_iso3, venue, result, gf, ga, competition}), wins, draws, losses, scraped_at',
     '11v11.com/teams/{slug}/tab/matches/ (Fase C). RLS enabled, sin policy pública.'),
    ('ia_h2h',
     'PK (team_a_code, team_b_code) con CHECK h2h_alphabetical (team_a < team_b).',
     'team_a_code, team_b_code, matches JSONB {total, gf_team_a, ga_team_a, source_team, source}, team_a_wins, team_b_wins, draws, last_played (DATE, null en 11v11 agregado), scraped_at',
     '11v11.com/teams/{slug}/tab/stats/ (Fase D.2). RLS enabled, sin policy pública. Dedup por pair antes de UPSERT (ON CONFLICT no admite misma fila 2x).'),
    ('ia_predictions',
     'PK match_id. Index idx_ia_predictions_computed (computed_at DESC).',
     'match_id, home_code, away_code, sign CHAR(1) CHECK (1|X|2), confidence SMALLINT 0-100, breakdown JSONB {elo_score, h2h_score, last5_score, raw_home_pct}, used_fallback BOOLEAN, computed_at',
     'EF compute (Fase E pendiente). **RLS + policy ia_predictions_public_read**: cualquier authenticated SELECT. Lectura desde frontend.'),
]
for t, pk, cols, origin in tables:
    row += 1
    for col, v in enumerate([t, pk, cols, origin], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin


# ── 2d. Bloque: Estado fases ─────────────────────────────────────────────────
row += 2
ws.cell(row=row, column=1).value = 'Estado Fases A-F (21 abr 2026)'
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = section_fill

row += 1
for col, h in enumerate(['Fase', 'Acción / Entregable', 'Commit / PR', 'Estado'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

phases = [
    ('A', 'Migración 4 tablas ia_* + EF porra-ia-compute esqueleto (router status/scrape_elo/scrape_last5/scrape_h2h/compute)',
     '968332a (PR #10)', '✅ Merged + migración aplicada'),
    ('B', 'scrape_elo via inside.fifa.com/api/ranking-overview',
     '4a32737 (PR #11)', '⚠️ Deprecada: solo datos hasta sept 2025'),
    ('B.2', 'scrape_elo via Wikipedia Module:SportsRankings/data/FIFA_World_Rankings (MediaWiki API + regex Lua)',
     'c845f3e (PR #12)', '✅ Merged + desplegada (v3). 211 países upserted.'),
    ('D', 'scrape_h2h via Wikipedia [País]_national_football_team_all-time_record',
     'cba5dcc (PR #13)', '⚠️ Deprecada: solo ~3/48 tienen página (ERR-24)'),
    ('D.2', 'scrape_h2h via 11v11.com/teams/{slug}/tab/stats/ (HTML regex, headers Chrome obligatorios)',
     'bbad657 (PR #14)', '✅ Merged + desplegada (v5). 815 pairs upserted (72% cobertura).'),
    ('C', 'scrape_last_n via 11v11.com/teams/{slug}/tab/matches/ (default N=8, limit parametrizable 1-20)',
     '5a87f1e (rama claude/fase-c-last-n, PR #15)', '🟡 EF v6 desplegada desde rama. PR abierto pendiente de merge (ERR-26).'),
    ('E', 'handleCompute: leer ELO + H2H + Racha, aplicar fórmula 50/25/25 con fallback 66/34, UPSERT ia_predictions',
     '—', '⏳ Pendiente'),
    ('F', 'Frontend scoring.js / ko.js: leer ia_predictions, pintar pronóstico en tarjeta, calcular bonus +1 IA-opuesta',
     '—', '⏳ Pendiente'),
]
for ph, act, commit, status in phases:
    row += 1
    for col, v in enumerate([ph, act, commit, status], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin
        if ph in ('B', 'D'):
            c.fill = PatternFill('solid', start_color='F2F2F2')
        elif ph == 'C':
            c.fill = PatternFill('solid', start_color='FFF2CC')


# ── 2e. Bloque: Estado contadores tablas ─────────────────────────────────────
row += 2
ws.cell(row=row, column=1).value = 'Estado tablas al 21 abr PM'
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = section_fill

row += 1
for col, h in enumerate(['Tabla', 'Filas', 'Última ingesta', 'Notas'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

counts = [
    ('ia_elo_fifa', '211', 'Wikipedia update 2026-04-01', 'Próximo refresh FIFA 9 jun 2026'),
    ('ia_h2h', '815', '21 abr 2026 (Fase D.2 smoke test)', 'Pares únicos entre mundialistas. 72% de 1.128 teóricos.'),
    ('ia_last5_results', '48', '21 abr 2026 (Fase C smoke test)', 'N=8 por selección; ARG 7 por caché 11v11. Total ~382 partidos en JSONB.'),
    ('ia_predictions', '0', '—', 'Poblada por Fase E pendiente.'),
]
for t, n, last, notes in counts:
    row += 1
    for col, v in enumerate([t, n, last, notes], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin


# ── 2f. Bloque: Headers obligatorios 11v11 ───────────────────────────────────
row += 2
ws.cell(row=row, column=1).value = 'Headers obligatorios 11v11.com (sin los 3 → 403, ver ERR-25)'
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = section_fill

row += 1
for col, h in enumerate(['Header', 'Valor', 'Rol', ''], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

headers_11v11 = [
    ('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36',
     'Chrome real. UA custom (ej. pm26-ia-predictor/1.0) devuelve 403.'),
    ('Accept', 'text/html,application/xhtml+xml', 'Explicita que esperamos HTML'),
    ('Accept-Language', 'en-US,en;q=0.9', 'Evita redirecciones de i18n del site'),
]
for h, v, role in headers_11v11:
    row += 1
    for col, val in enumerate([h, v, role, ''], 1):
        c = ws.cell(row=row, column=col); c.value = val
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin


# ── Anchos ───────────────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 55


# ── 3. Sheet Errores: renombrar a ERR-01..ERR-26 y añadir 24/25/26 ──────────
old_name = next((n for n in wb.sheetnames if n.startswith('Errores ERR-01')), None)
if old_name and old_name != 'Errores ERR-01..ERR-26':
    ws_err = wb[old_name]
    ws_err.title = 'Errores ERR-01..ERR-26'
else:
    ws_err = wb['Errores ERR-01..ERR-26']

# Añadir 24/25/26 si no existen
existing_ids = {
    str(ws_err.cell(row=r, column=1).value or '').strip()
    for r in range(1, ws_err.max_row + 1)
}
new_errors = [
    ('ERR-24',
     'Wikipedia sin cobertura masiva de páginas "[País]_national_football_team_all-time_record"',
     'Solo ~3/48 selecciones tienen esa página. Encabezados y formato wikitext inconsistentes entre las que sí existen.',
     'Migrar a 11v11.com/stats (Fase D.2 commit bbad657). Validar ≥5 muestras heterogéneas antes de elegir una fuente para scraping masivo.'),
    ('ERR-25',
     'fetch a 11v11.com devuelve HTTP 403 sin 3 headers obligatorios',
     'Anti-bot básico: exige User-Agent de Chrome real + Accept: text/html + Accept-Language: en-US. Faltar cualquiera = 403.',
     'Constante fetchHeaders top-level en supabase/functions/porra-ia-compute/index.ts con los 3 headers. Aplicado en handleScrapeH2h y handleScrapeLast5.'),
    ('ERR-26',
     'pg_net no soporta HTTP PUT: no se puede mergear PR desde Supabase',
     'pg_net solo expone net.http_get / net.http_post / net.http_delete. Merge de GitHub requiere PUT /repos/:owner/:repo/pulls/:n/merge.',
     'Workaround: deploy directo del código con deploy_edge_function (evita PUT). PR se mergea después desde MCP GitHub / UI / gh CLI.'),
]
for eid, sym, cause, fix in new_errors:
    if eid in existing_ids:
        continue
    nr = ws_err.max_row + 1
    for col, v in enumerate([eid, sym, cause, fix], 1):
        c = ws_err.cell(row=nr, column=col)
        c.value = v
        c.font = cell_font
        c.alignment = cell_align
        c.border = border_thin
        c.fill = highlight_fill

# ── Footer global ────────────────────────────────────────────────────────────
for nm in wb.sheetnames:
    wb[nm].oddFooter.center.text = '&B Esquema Porra Mundial 2026 &R Actualizado: 21 abr 2026 (IA Predictor A-C)'

wb.save(DST)
print(f'XLSX guardado: {DST}')
print(f'Sheets: {wb.sheetnames}')
