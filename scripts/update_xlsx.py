# update_xlsx.py — regenera ESQUEMA_SISTEMA_PORRA2026.xlsx con las hojas nuevas
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = 'ESQUEMA_SISTEMA_PORRA2026.xlsx'
DST = 'ESQUEMA_SISTEMA_PORRA2026.xlsx'

wb = load_workbook(SRC)

# Estilos
title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill('solid', start_color='1F4E78')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='2E75B6')
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
border_thin = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

# ── 1. Edge Functions: actualizar versiones + añadir create-league v2 ──
ws_ef = wb['Edge Functions']
for row in range(1, ws_ef.max_row + 1):
    a = str(ws_ef.cell(row=row, column=1).value or '')
    if 'porra-match-live' in a:
        ws_ef.cell(row=row, column=1).value = 'porra-match-live v13'
    elif 'porra-apify-webhook' in a:
        ws_ef.cell(row=row, column=1).value = 'porra-apify-webhook v7'
    elif 'porra-fix-encoding' in a:
        ws_ef.cell(row=row, column=1).value = 'porra-fix-encoding v6'
    elif 'update-results' in a:
        ws_ef.cell(row=row, column=1).value = 'update-results v4'

has_cl = any('create-league' in str(ws_ef.cell(row=r, column=1).value or '') for r in range(1, ws_ef.max_row + 1))
if not has_cl:
    nr = ws_ef.max_row + 1
    vals = ['create-league v2', 'Frontend cuando usuario no-admin crea porra',
            'Supabase DB (leagues, league_members)',
            "Permite a cualquier usuario crear una porra. Límite: 3 porras activas para no-admin. verify_jwt=false por JWT ES256 (ERR-16) — valida manualmente.",
            '✅ En producción']
    for col, v in enumerate(vals, 1):
        c = ws_ef.cell(row=nr, column=col)
        c.value = v
        c.font = cell_font
        c.alignment = cell_align
        c.border = border_thin

# ── 2. APIs + DB: actualizar league_members con groups_saved ──
ws_db = wb['APIs + DB']
lm_found = False
for row in range(1, ws_db.max_row + 1):
    if str(ws_db.cell(row=row, column=1).value or '').strip() == 'league_members':
        lm_found = True
        cur = str(ws_db.cell(row=row, column=4).value or '')
        if 'groups_saved' not in cur:
            ws_db.cell(row=row, column=4).value = (cur +
                ". NUEVO 19abr26: groups_saved JSONB DEFAULT '{}' — mapa {A:true, B:true...} " +
                "que refleja qué grupos guardó el usuario en vista móvil. Policy lm_update requiere porra_cerrada=false.")
        break
if not lm_found:
    nr = ws_db.max_row + 1
    vals = ['league_members', 'Frontend (usuario autenticado)', 'Supabase Postgres',
            "Membresías de usuarios en ligas. Columnas: league_id, user_id, porra_cerrada, cerrada_at. NUEVO 19abr26: groups_saved JSONB DEFAULT '{}' — mapa {A:true...}. Policy lm_update requiere porra_cerrada=false.",
            'RLS Supabase', '✅ Activa']
    for col, v in enumerate(vals, 1):
        c = ws_db.cell(row=nr, column=col)
        c.value = v
        c.font = cell_font
        c.alignment = cell_align
        c.border = border_thin

# ── 3. Hoja Frontend Mobile ──
if 'Frontend Mobile' in wb.sheetnames:
    del wb['Frontend Mobile']
ws = wb.create_sheet('Frontend Mobile')
ws['A1'] = 'PORRA MUNDIAL 2026 — Frontend móvil (feat/mobile-grupos-focus merged 19abr26)'
ws.merge_cells('A1:D1')
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center_align

row = 3
ws.cell(row=row, column=1).value = 'MÓDULOS JS — public/js/'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = PatternFill('solid', start_color='D9E1F2')

row += 1
for col, h in enumerate(['Módulo', 'Líneas', 'Rol', 'Notas'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

modules = [
    ('main-entry.js', '~20', 'Entry point Vite type=module', 'Bootstrap clásico — dispara carga ordenada'),
    ('data.js', '215', 'Datos torneo + estado global', 'EQUIPOS, PARTIDOS, PHRASES_GRUPO, boostPicks, groupSaved'),
    ('scoring.js', '1184', 'Motor puntos + tarjetas + premios', 'renderMatchCard, updateCardUI, AWARDS_CFG'),
    ('ui-groups.js', '167', 'Grupos + vista Jornada', 'renderGroups, tarjetas compactas jornada'),
    ('ui-groups-mobile.js', '~700', 'Acordeón + focus layer + carrusel', 'NUEVO 19abr26. ensureFocusLayer, openMobileFocus, smart boost, slide 7, guardar async'),
    ('ko.js', '1048', 'Bracket KO + IA', 'Fase eliminatoria'),
    ('ui-nav.js', '653', 'SPA nav + modal + welcome', 'showPage'),
    ('auth.js', '~400', 'Auth Supabase', 'NUEVO 19abr26: 4ª query Promise.all a league_members.groups_saved'),
    ('leagues.js', '~300', 'Ligas', 'Multi-liga, código invitación'),
    ('misc.js', '~200', 'Utils UI (paralelo)', 'showToast, lockUI'),
    ('scoreboard.js', '~250', 'Clasificación', 'Ranking porras'),
    ('close-porra.js', '~100', 'Cierre pronósticos', 'Bloqueo al cerrar'),
    ('admin.js', '~500', 'Panel admin + dados', 'dice.js integrado'),
    ('bracket-results.js', '~300', 'Vista resultados KO', 'Mobile min-width 260px'),
    ('live-sync.js', '~150', 'Sync live scores', 'Realtime Supabase'),
    ('ui-directo.js', '~200', 'Vista Directo', 'Panel live partidos simultáneos'),
]
for mod, lns, role, notes in modules:
    row += 1
    for col, v in enumerate([mod, lns, role, notes], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin
        if 'NUEVO 19abr26' in str(v):
            c.fill = PatternFill('solid', start_color='E2EFDA')

row += 2
ws.cell(row=row, column=1).value = 'CSS público — public/css/ (7 ficheros linkeados desde index.html tras refactor 9e93fe8)'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = PatternFill('solid', start_color='D9E1F2')

row += 1
for col, h in enumerate(['Fichero', 'Tamaño aprox', 'Rol', 'Notas'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

css_files = [
    ('base.css', '67 KB', 'Base + acordeón móvil + focus layer', 'mobile-collapsed, mobile-focus-layer (fuera @media), smart boost'),
    ('welcome.css', '22 KB', 'Pantalla splash/welcome', 'Hero + scroll-cue'),
    ('ko.css', '64 KB', 'Fase eliminatoria + awards', 'Bracket KO, cards premios individuales'),
    ('admin.css', '57 KB', 'Panel administrador', 'Dados, simulador, simulacros'),
    ('bracket-results.css', '~15 KB', 'Vista resultados KO', 'Mobile: min-width 260px columna activa'),
    ('boost.css', '~10 KB', 'Boost x2 fuego canvas', 'Ticker boosts, efecto Canvas'),
    ('directo.css', '~12 KB', 'Vista Directo live', 'Panel en vivo, pills CEST'),
]
for f, sz, role, notes in css_files:
    row += 1
    for col, v in enumerate([f, sz, role, notes], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin

row += 2
ws.cell(row=row, column=1).value = 'Rediseño móvil grupos — componentes (PR#9 → refactor CSS 9e93fe8)'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).fill = PatternFill('solid', start_color='D9E1F2')

row += 1
for col, h in enumerate(['Componente', 'Activación', 'Descripción', 'Persistencia'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

comps = [
    ('Acordeón colapsado', 'Viewport ≤640px', 'Grupos plegados con header + barra progreso', 'Estado memoria'),
    ('Focus layer pantalla completa', 'Tap en header grupo', 'position:fixed inset:0 z-index:100. 7 slides', '__mobileFocusState'),
    ('Carrusel tarjetas', 'Dentro focus layer', '6 tarjetas + slide 7 summary. Swipe threshold 50px. Dots + flechas', '__mobileFocusState.slide'),
    ('Smart boost multi-jornada', 'Botón boost', 'Detecta conflictos inter-grupo mismo match_date. confirm() si conflicto', 'boost_picks tabla'),
    ('Slide 7 summary', 'Slide final', 'Mueve #gtable-L al slide. Fallback si falta tabla', 'Tabla del grupo'),
    ('Guardar/deshacer', '💾 Guardar grupo', '4 estados: guardar/guardado/deshacer/bloqueado. canSaveGroup verifica 6 partidos', 'league_members.groups_saved JSONB'),
    ('Barra progreso grupo', 'Header colapsado', 'Cuenta partidos con pred. completa', 'Estado memoria'),
    ('Frase motivacional', 'Header focus', 'PHRASES_GRUPO[letra][percent] varía según completitud', 'data.js'),
]
for comp, trig, desc, pers in comps:
    row += 1
    for col, v in enumerate([comp, trig, desc, pers], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 55

# ── 4. Hoja Errores ERR-01..ERR-22 ──
if 'Errores ERR-01..ERR-22' in wb.sheetnames:
    del wb['Errores ERR-01..ERR-22']
ws = wb.create_sheet('Errores ERR-01..ERR-22')
ws['A1'] = 'PORRA MUNDIAL 2026 — Inventario de errores conocidos'
ws.merge_cells('A1:D1')
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center_align

row = 3
for col, h in enumerate(['ID', 'Síntoma', 'Causa raíz', 'Fix'], 1):
    c = ws.cell(row=row, column=col); c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = center_align; c.border = border_thin

errors = [
    ('ERR-01', 'DOMContentLoaded no dispara en classic scripts async', 'Script ya cargado cuando registras listener', 'if (document.readyState===loading) addEventListener else run()'),
    ('ERR-02', 'const globals no accesibles entre módulos', 'const no se expone en window automáticamente', 'window.X = X explícito'),
    ('ERR-03', 'Vite no copia assets de /css en build', 'Vite solo copia public/', 'Mover css/ → public/css/'),
    ('ERR-04', 'Whitespace en secrets Vault rompe HTTP', 'Secrets con newline al final', 'trim() antes de usar'),
    ('ERR-05', 'pg_net timeout >30s', 'Timeout fijo ~30s', 'Async + webhook pattern'),
    ('ERR-06', 'vercel.json rompe MIME ES modules', 'Wildcard source sobrescribe Content-Type', 'NO crear vercel.json'),
    ('ERR-07', 'GitHub raw bloqueado proxy Claude.ai', 'Dominio no whitelisted', 'net.http_post + GITHUB_TOKEN desde Supabase MCP'),
    ('ERR-08', 'SofaScore API 403 (Cloudflare)', 'Bot Management detecta peticiones no-browser', 'Playwright proxy residencial + page.evaluate(fetch)'),
    ('ERR-09', 'checkIsAdmin async no detecta admin render inicial', 'Race condition render vs query profiles', 'Retries + re-render al completar'),
    ('ERR-10', 'Webhook Apify no llega', 'Filtro eventTypes incorrecto', 'ACTOR.RUN.SUCCEEDED + FAILED'),
    ('ERR-11', 'Twilio 21211 invalid To', 'Falta prefijo whatsapp:', 'whatsapp:+34XXX'),
    ('ERR-12', 'porra-whatsapp-send pg_net 400', 'pg_net no soporta form-urlencoded', 'fetch() Content-Type x-www-form-urlencoded'),
    ('ERR-13', 'porra-fix-encoding 404 intermitente', 'Reuso conexión EF antigua', 'Retry 5-10s'),
    ('ERR-14', 'Simulacro admin no aparece al cargar', 'checkIsAdmin async no completa antes render', 'Retry 3x + re-render'),
    ('ERR-15', 'QA sobrescribir encrypted_password invalida sesiones', 'Manipular auth.users rompe Supabase Auth', 'generateLink para reset'),
    ('ERR-16', 'EF verify_jwt=true falla con JWT ES256', 'Supabase no valida ES256 automáticamente', 'verify_jwt=false + jose manual'),
    ('ERR-17', 'net.http_put no existe en pg_net', 'pg_net solo GET/POST/DELETE', 'Chrome MCP para merge PR'),
    ('ERR-18', 'css/ no servido prod tras merge', 'Vite no copia /css/ del root', 'Mover a public/css/. Parcial — real era ERR-22'),
    ('ERR-19', 'openMobileFocus sin try/catch deja body inconsistente', 'body.overflow=hidden antes de confirmar éxito', 'try/catch + overflow al final (después eliminado)'),
    ('ERR-20', 'body.overflow=hidden bloquea scroll iPhone Safari', 'WebKit iOS quirk persistente', 'Eliminar uso de body.overflow=hidden. position:fixed basta'),
    ('ERR-21', '.mobile-focus-layer inline en viewport >640px', 'Reglas CSS dentro @media', 'Sacar base del @media + visibility:hidden/visible'),
    ('ERR-22', 'Reglas CSS nuevas NO aplican prod (root cause real)', 'index.html tenía <style> inline con "Archivo destino:X.css" nunca migrados. <link> no existían', 'Refactor commit 9e93fe8: extraer <style> → public/css/ + 4 <link> en <head>. getComputedStyle() test definitivo'),
]
for eid, sym, cause, fix in errors:
    row += 1
    for col, v in enumerate([eid, sym, cause, fix], 1):
        c = ws.cell(row=row, column=col); c.value = v
        c.font = cell_font; c.alignment = cell_align; c.border = border_thin
        if eid in ('ERR-18', 'ERR-19', 'ERR-20', 'ERR-21', 'ERR-22'):
            c.fill = PatternFill('solid', start_color='FFF2CC')

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 55

# Footer
for nm in wb.sheetnames:
    wb[nm].oddFooter.center.text = '&B Esquema Porra Mundial 2026 &R Actualizado: 20 abr 2026 (commit 0d3d636)'

wb.save(DST)
print(f'XLSX guardado: {DST}')
print(f'Sheets: {wb.sheetnames}')
